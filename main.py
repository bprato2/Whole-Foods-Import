import os
from receiveinput import process_html_file
import pandas as pd
import shutil
import openpyxl

folder_path = "C:/Users/blair/OneDrive/Documents/GitHub/Whole-Foods-Import/Whole Foods PDFs"

orders_compiled = []
for filename in os.listdir(folder_path):
    file_path = os.path.join(folder_path, filename)
    if os.path.isfile(file_path):
        sales_line = process_html_file(file_path)
        orders_compiled.append(sales_line)

# Read the Excel file into a DataFrame
df = pd.read_excel("Whole Foods Stores.xlsx")


def add_lists_together(data_list):
    sales_header = {
        'Document Type': [],
        'No.': [],
        'Sell-to Customer No.': [],
        'Bill-to Customer No.': [],
        'Ship-to Code': [],
        'Order Date': [],
        'Due Date': [],
        'Location Code': [],
        'Customer Posting Group': [],
        'Customer Price Group': [],
        'Gen. Bus. Posting Group': [],
        'Sell-to Customer Name': [],
        'Sell-to Customer Name 2': [],
        'Sell-to Address': [],
        'Sell-to Address 2': [],
        'Sell-to City': [],
        'Sell-to ZIP code': [],
        'External Document No.': [],
        'Delivery Info': [],
        'Delivery Exception': [],
        'Route': [],
        'Store Delivery Location': []
    }

    sales_line = {
        'Document Type': [],
        'Document No.': [],
        'Line No.': [],
        'Type': [],
        'No.': [],
        'Location Code': [],
        'Shipment Date': [],
        'Description': [],
        'Quantity': [],
        'Unit Price': []
    }

    for data in data_list:
        sales_line['Document Type'] += data['Document Type']
        sales_line['Document No.'] += data['Document No.']
        sales_line['Line No.'] += data['Line No.']
        sales_line['Type'] += data['Type']
        sales_line['No.'] += data['No.']
        sales_line['Location Code'] += data['Location Code']
        sales_line['Shipment Date'] += data['Shipment Date']
        sales_line['Description'] += data['Description']
        sales_line['Quantity'] += data['Quantity']
        sales_line['Unit Price'] += data['Unit Price']
        customer_number = (data['Customer Number'])
        sales_header['Document Type'] += ['Order']
        sales_header['No.'] += [data['Document No.'][0]]
        sales_header['Sell-to Customer No.'] += [customer_number]
        sales_header['Bill-to Customer No.'] += ['WHOLE FOODS CORPORAT']
        sales_header['Ship-to Code'] += [df.loc[df['No.'] == customer_number, "Ship-to Code"].values[0]]
        sales_header['Order Date'] += [data['Order Date']]
        sales_header['Due Date'] += ['']
        sales_header['Location Code'] += ['ELIZABETH']
        sales_header['Customer Posting Group'] += ['WHOLE FOODS']
        sales_header['Customer Price Group'] += ['WHOLE FOOD']
        sales_header['Gen. Bus. Posting Group'] += ['ALL']
        sales_header['Sell-to Customer Name'] += [df.loc[df['No.'] == customer_number, "Sell-to Customer Name"].values[0]]
        sales_header['Sell-to Customer Name 2'] += ['']
        sales_header['Sell-to Address'] += [df.loc[df['No.'] == customer_number, 'Sell-to Address'].values[0]]
        sales_header['Sell-to Address 2'] += [df.loc[df['No.'] == customer_number, 'Sell-to Address 2'].values[0]]
        sales_header['Sell-to City'] += [df.loc[df['No.'] == customer_number, 'Sell-to City'].values[0]]
        sales_header['Sell-to ZIP code'] += [df.loc[df['No.'] == customer_number, 'Sell-to ZIP Code'].values[0]]
        sales_header['External Document No.'] += [data['Order Number']]
        sales_header['Delivery Info'] += [df.loc[df['No.'] == customer_number, 'Delivery Info'].values[0]]
        sales_header['Delivery Exception'] += [df.loc[df['No.'] == customer_number, 'Delivery Exception'].values[0]]
        sales_header['Route'] += [df.loc[df['No.'] == customer_number, 'Route'].values[0]]
        sales_header['Store Delivery Location'] += [df.loc[df['No.'] == customer_number, 'Store Delivery Location'].values[0]]

    return sales_line, sales_header


sales_line_data, sales_header_data = add_lists_together(orders_compiled)

'''for key, value in sales_header_data.items():
    print(key, len(value))'''

'''# Create two DataFrames
sales_line_df = pd.DataFrame(sales_line_data)
sales_header_df = pd.DataFrame(sales_header_data)'''

# Write each DataFrame to its own sheet in the same Excel file
'''with pd.ExcelWriter('orders.xlsx') as writer:
    sales_line_sheet = writer.book.create_sheet("Sales Line")

    sales_line_sheet['A1'] = "WHOLE FOODS IMPORT"
    sales_line_sheet['B1'] = "Sales Line"
    sales_line_sheet['C1'] = "37"

    # Write the data in sales_line_df starting from cell A3
    start_row = 2
    sales_line_df.to_excel(writer, sheet_name="Sales Line", startrow=start_row, index=False)

    sales_header_sheet = writer.book.create_sheet("Sales Header")

    sales_header_sheet['A1'] = "WHOLE FOODS IMPORT"
    sales_header_sheet['B1'] = "Sales Header"
    sales_header_sheet['C1'] = "36"

    # Write the data in sales_header_df starting from cell A3
    start_row = 2
    sales_header_df.to_excel(writer, sheet_name="Sales Header", startrow=start_row, index=False)
    # Close the writer
    writer.save()



# Load the workbook
wb = openpyxl.load_workbook('orders.xlsx')

# Define the XLM map path
xlm_map_path = 'header xml file.xml'

# Add the XLM map to the workbook
wb.vba_project.import_file(xlm_map_path)

# Save the changes to the workbook
wb.save('orders.xlsx')'''


# Load the existing workbook
# workbook = openpyxl.load_workbook("Sales Header and Line File.xlsx")

# Create two DataFrames
sales_line_df = pd.DataFrame(sales_line_data)
sales_header_df = pd.DataFrame(sales_header_data)

# Write each DataFrame to its own sheet in the same Excel file
with pd.ExcelWriter('orders.xlsx', engine='openpyxl', mode='w') as writer:
    workbook = writer.book

    # Create another sheet in the workbook
    sales_header_sheet = workbook.create_sheet("Sales Header")

    # Write headers
    sales_header_sheet['A1'] = "WHOLE FOODS IMPORT"
    sales_header_sheet['B1'] = "Sales Header"
    sales_header_sheet['C1'] = "36"

    # Write the data in sales_header_df starting from cell A3
    start_row = 3
    sales_header_df.to_excel(writer, sheet_name="Sales Header", startrow=start_row, index=False)

    # Create a new sheet in the workbook
    sales_line_sheet = workbook.create_sheet("Sales Line")

    # Write the values in cells A1:C1
    sales_line_sheet['A1'] = "WHOLE FOODS IMPORT"
    sales_line_sheet['B1'] = "Sales Line"
    sales_line_sheet['C1'] = "37"

    # Write the data in sales_line_df starting from cell A3
    start_row = 3
    sales_line_df.to_excel(writer, sheet_name="Sales Line", startrow=start_row, index=False)


# Save the changes to the Excel file
